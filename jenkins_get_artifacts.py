import datetime
import openpyxl
import pytz
from jenkinsapi.jenkins import Jenkins

import var

jenkins_user = var.jenkins_user
jenkins_token = var.jenkins_token
jenkins_url = var.jenkins_url
job_name = var.job_name

builds_list_success = []
builds_list_failure = []
builds_list_aborted = []
builds_list_process = []

colection_success = {}
colection_failure = {}
colection_aborted = {}
colection_process = {}

artifacts = []

time_start = datetime.datetime(2021, 10, 1, tzinfo=pytz.UTC)
time_end = datetime.datetime.now(tz=pytz.UTC)

excel_path = var.excel_path  # path to excel (.xlsx will be create)
artifacts_name = var.artifacts_name  # list of artifacts name


def get_server(url):
    try:
        jenkins = Jenkins(baseurl=url, username=jenkins_user, password=jenkins_token, timeout=100)
        print(f"Jenkins version: {jenkins.version}")
        return jenkins
    except Exception:
        print(f"Error connect to Jenkins. Use VPN.")
        exit()


def get_builds_list(server, job_name):
    print(f"Downloading builds list.")
    job = server[job_name]
    builds = job.get_build_dict()
    builds_list = list(builds.keys())

    print(f"\nCollecting data "
          f"from {time_start.strftime('%d/%m/%y %H:%M')} "
          f"to {time_end.strftime('%d/%m/%y %H:%M')}.")
    print(f"List builds from {job.get_full_name()} job: {builds_list}")
    print(f"List has {len(builds_list)} builds")
    print(f"Sorting builds list")

    tmp = 1
    for build_id in builds_list:
        build_time = job.get_build(build_id).get_timestamp()
        build_status = job.get_build(build_id).get_status()

        print(f"build #{build_id} \t"
              f"time: {build_time.strftime('%d/%m/%y %H:%M')} \t"
              f"status: {build_status} \t"
              f"{tmp}/{len(builds_list)}")

        if time_end >= build_time >= time_start:
            if build_status is None:
                builds_list_process.append(build_id)
                colection_process.update({build_id: {"time": build_time}})
            if build_status == "ABORTED":
                builds_list_aborted.append(build_id)
                colection_aborted.update({build_id: {"time": build_time}})
            if build_status == "FAILURE":
                builds_list_failure.append(build_id)
                colection_failure.update({build_id: {"time": build_time}})
            if build_status == "SUCCESS":
                builds_list_success.append(build_id)
                colection_success.update({build_id: {"time": build_time}})
        else:
            break

        tmp = tmp + 1

    print("\nBUILD DOWNLOAD RESULT :")
    print(f"SUCCESS: \t {builds_list_success}")
    print(f"FAILURE: \t {builds_list_failure}")
    print(f"PROCESS: \t {builds_list_process}")
    print(f"ABORTED: \t {builds_list_aborted}")


def get_artifacts(server, job_name, data):
    print(f"Downloading artifacts from {job_name}.")
    for job in server.get_jobs():
        if job[0] == job_name:
            job_instance = server.get_job(job[0])
            tmp = 1
            for build_id in data.keys():
                try:
                    print(f"Downloading artifacts from {job_name}/#{build_id} - start\t{tmp}/{len(data.keys())}")
                    build = job_instance.get_build(buildnumber=build_id)
                    data[build_id].update({"artifact": {}})
                    for artifact in build.get_artifacts():
                        artifact_name = artifact.filename
                        if artifact_name[0:4] == artifacts_name[0][0:4] \
                                and artifact_name[-3:] == artifacts_name[0][-3:]:
                            artifact_name = artifacts_name[0]
                        if len(artifact.get_data()) == 0:
                            data[build_id]["artifact"].update({artifact_name: 0})
                        else:
                            data[build_id]["artifact"].update({artifact_name: 1})
                    print(f"Downloading artifacts from {job_name}/#{build_id} - done")
                except Exception:
                    print(f"Downloading artifacts from {job_name}/#{build_id} - error")
                tmp = tmp + 1
    return data


def analize(colection):
    print("\nNumber of builds:")
    print(f"Success builds:\t {len(builds_list_success)}")
    print(f"Failed builds:\t {len(builds_list_failure)}")
    print(f"Aborted builds:\t {len(builds_list_aborted)}")
    print(f"Process builds:\t {len(builds_list_process)}")
    print(f"\nNumber of empty files in artifacts "
          f"from {time_start.strftime('%d/%m/%y %H:%M')} "
          f"to {time_end.strftime('%d/%m/%y %H:%M')}")

    for artifact in artifacts_name:
        amount = 0
        for build in list(colection.keys()):
            if colection[build]["artifact"].get(artifact) == 0:
                amount = amount + 1
        print(f"{amount}\tempty artifacts {artifact}")


def to_excel(colection, build_list, sheet_name, workbook):
    print(f"\nCreating excel for {sheet_name}")
    sheet = workbook.create_sheet(sheet_name)
    sheet.cell(row=1, column=1).value = "build"
    sheet.cell(row=1, column=2).value = "date"
    for column in range(0, len(artifacts_name)):
        sheet.cell(row=1, column=column + 3).value = artifacts_name[column]
    for row in range(2, len(build_list) + 2):
        build_id = list(colection.keys())[row - 2]
        build_time = colection[build_id]["time"]

        sheet.cell(row=row, column=1).value = build_id
        sheet.cell(row=row, column=2).value = str(build_time.strftime('%d/%m/%y'))

        for value in range(0, len(artifacts_name) + 1):
            try:
                sheet.cell(row=row, column=value + 3).value = colection[build_id]["artifact"][
                    artifacts_name[value]]
            except Exception:
                sheet.cell(row=row, column=value + 3).value = ""

    return workbook


if __name__ == '__main__':
    jenkins_server = get_server(jenkins_url)

    get_builds_list(jenkins_server, job_name)

    colection_success = get_artifacts(jenkins_server, job_name, colection_success)
    colection_failure = get_artifacts(jenkins_server, job_name, colection_failure)
    colection_aborted = get_artifacts(jenkins_server, job_name, colection_aborted)

    analize(colection=colection_success)
    workbook = openpyxl.Workbook()
    workbook = to_excel(colection=colection_failure,
                        build_list=builds_list_failure,
                        sheet_name="failed builds",
                        workbook=workbook)

    workbook = to_excel(colection=colection_success,
                        build_list=builds_list_success,
                        sheet_name="success builds",
                        workbook=workbook)

    workbook = to_excel(colection=colection_aborted,
                        build_list=builds_list_aborted,
                        sheet_name="aborted builds",
                        workbook=workbook)

    workbook.save(excel_path)
